from pathlib import Path

from openpyxl import Workbook

from mmi.ingest.excel import ExcelAdapter


def test_excel_extract_headers_and_row_anchors(tmp_path: Path) -> None:
    path = tmp_path / "historial.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Fallos"
    ws.append(["asset_id", "fecha", "codigo"])
    ws.append(["P-101", "2024-01-10", "SEAL-01"])
    ws.append([None, None, None])
    ws.append(["P-102", None, "BRG-02"])
    wb.save(path)
    wb.close()

    doc = ExcelAdapter().extract(path)
    assert doc.quality == "pass"
    assert len(doc.records) == 2
    assert doc.records[0].sheet == "Fallos"
    assert doc.records[0].row == 2
    assert doc.records[0].values["asset_id"] == "P-101"
    assert doc.records[1].values["fecha"] is None
    assert doc.anchors[0].row == 2
    assert "| _excel_row_ |" in doc.markdown


def test_excel_template_only_is_review_or_reject(tmp_path: Path) -> None:
    path = tmp_path / "plantilla.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["modo_falla", "efecto", "causa"])
    wb.save(path)
    wb.close()

    doc = ExcelAdapter().extract(path)
    assert doc.records == []
    assert doc.quality in {"review", "reject"}
    assert doc.sheets[0].status == "template_only"


def test_excel_merged_group_headers(tmp_path: Path) -> None:
    path = tmp_path / "checklist.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["Grupo A", "", "Grupo B", ""])
    ws.append(["Item", "Peso", "Item", "Peso"])
    ws.append(["Accesibilidad", "2", "Mantenibilidad", "3"])
    wb.save(path)
    wb.close()

    doc = ExcelAdapter().extract(path)
    assert doc.quality == "pass"
    cols = doc.sheets[0].columns
    assert any("Grupo A" in c for c in cols)
    assert doc.records[0].row == 3
