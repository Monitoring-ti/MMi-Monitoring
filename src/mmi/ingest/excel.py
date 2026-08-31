"""Extractor Excel industrial: cabeceras multi-nivel, nulls, anclas sheet/row."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from mmi.ingest.ports import (
    CitationAnchor,
    ExtractedDocument,
    SheetSummary,
    SpreadsheetPort,
    SpreadsheetRecord,
)


class ExcelAdapter(SpreadsheetPort):
    """Normaliza un workbook a registros citables + Markdown por hoja."""

    def extract(self, path: Path) -> ExtractedDocument:
        wb = load_workbook(path, data_only=True, read_only=True)
        notes: list[str] = []
        anchors: list[CitationAnchor] = []
        records: list[SpreadsheetRecord] = []
        sheets: list[SheetSummary] = []
        parts: list[str] = [f"# {path.name}", ""]

        for ws in wb.worksheets:
            rows = [[_cell_raw(c) for c in r] for r in ws.iter_rows(values_only=True)]
            if not rows or not any(any(c for c in row) for row in rows):
                notes.append(f"Hoja vacía: {ws.title}")
                sheets.append(
                    SheetSummary(
                        name=ws.title,
                        header_row=None,
                        data_rows=0,
                        columns=[],
                        status="empty",
                    )
                )
                continue

            located = _locate_table(rows)
            if located is None:
                notes.append(f"Sin header detectado: {ws.title}")
                sheets.append(
                    SheetSummary(
                        name=ws.title,
                        header_row=None,
                        data_rows=0,
                        columns=[],
                        status="no_header",
                    )
                )
                parts.append(f"## {ws.title}\n\n_needs_review: sin encabezados_\n")
                continue

            header_idx, data_start, headers = located
            keep = _used_columns(headers, rows, data_start)
            headers_kept = [headers[c] or f"col_{c + 1}" for c in keep]

            sheet_records: list[SpreadsheetRecord] = []
            for r_idx in range(data_start, len(rows)):
                raw = rows[r_idx]
                if not any(raw):
                    continue
                values: dict[str, str | None] = {}
                parts_line: list[str] = []
                for c, label in zip(keep, headers_kept):
                    val = raw[c] if c < len(raw) else ""
                    shown: str | None = val if val else None
                    values[label] = shown
                    parts_line.append(f"{label}: {shown if shown is not None else 'null'}")
                excel_row = r_idx + 1
                text_line = " | ".join(parts_line)
                rec = SpreadsheetRecord(
                    sheet=ws.title,
                    row=excel_row,
                    values=values,
                    text_line=text_line,
                )
                sheet_records.append(rec)
                records.append(rec)
                anchors.append(CitationAnchor(sheet=ws.title, row=excel_row))

            if not sheet_records:
                status = "template_only"
                notes.append(f"Solo cabecera (sin filas de datos): {ws.title}")
            else:
                status = "ok"

            sheets.append(
                SheetSummary(
                    name=ws.title,
                    header_row=header_idx + 1,
                    data_rows=len(sheet_records),
                    columns=headers_kept,
                    status=status,
                )
            )

            parts.append(f"## {ws.title}")
            parts.append("")
            parts.append(f"_header_row: {header_idx + 1} · filas: {len(sheet_records)}_")
            parts.append("")
            if sheet_records:
                parts.append("| _excel_row_ | " + " | ".join(_md(h) for h in headers_kept) + " |")
                parts.append("| --- | " + " | ".join("---" for _ in headers_kept) + " |")
                # Preview acotado en markdown; el JSON lleva todas las filas
                preview_n = min(50, len(sheet_records))
                for rec in sheet_records[:preview_n]:
                    cells = [_md(rec.values.get(h) or "") for h in headers_kept]
                    parts.append(f"| {rec.row} | " + " | ".join(cells) + " |")
                if len(sheet_records) > preview_n:
                    parts.append(
                        f"\n_… {len(sheet_records) - preview_n} filas más "
                        f"(ver JSON / visor HTML)_\n"
                    )
            else:
                parts.append("_sin filas de datos_\n")
            parts.append("")

        wb.close()

        data_rows = sum(s.data_rows for s in sheets)
        if data_rows == 0 and not any(s.status == "ok" for s in sheets):
            quality = "reject"
        elif notes:
            quality = "review"
        else:
            quality = "pass"

        return ExtractedDocument(
            markdown="\n".join(parts).strip() + "\n",
            quality=quality,
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            source_path=str(path),
            anchors=anchors,
            notes=notes,
            records=records,
            sheets=sheets,
            meta={
                "file_name": path.name,
                "sheet_count": len(sheets),
                "record_count": len(records),
                "quality": quality,
            },
        )


def _cell_raw(v: object) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _md(value: str) -> str:
    return value.replace("|", "\\|").replace("\n", " ").strip()


def _filled(row: list[str]) -> int:
    return sum(1 for c in row if c)


def _is_subheader(row: list[str]) -> bool:
    vals = [c for c in row if c]
    if len(vals) < 2:
        return False
    short = sum(1 for v in vals if len(v) <= 4)
    return short / len(vals) >= 0.6


def _forward_fill(row: list[str]) -> list[str]:
    out: list[str] = []
    last = ""
    for c in row:
        if c:
            last = c
        out.append(last)
    return out


def _locate_table(rows: list[list[str]]) -> tuple[int, int, list[str]] | None:
    """Devuelve (header_idx 0-based, data_start, headers) o None."""
    n = len(rows)
    best_header, best_score = None, 0
    scan = min(n, 40)
    for i in range(scan):
        score = _filled(rows[i])
        if score > best_score:
            best_header, best_score = i, score
    if best_header is None or best_score < 2:
        return None

    headers = _build_headers(rows, best_header)
    nxt = best_header + 1
    if nxt < n and _is_subheader(rows[nxt]):
        data_start = best_header + 2
    else:
        data_start = best_header + 1
    return best_header, data_start, headers


def _build_headers(rows: list[list[str]], header_idx: int) -> list[str]:
    main = _forward_fill(rows[header_idx])
    if header_idx > 0 and _filled(rows[header_idx - 1]) >= 2:
        top = _forward_fill(rows[header_idx - 1])
        width = max(len(top), len(main))
        top += [""] * (width - len(top))
        main += [""] * (width - len(main))
        main = [
            f"{t} > {m}".strip(" >") if t and t != m else (m or t)
            for t, m in zip(top, main)
        ]
    nxt = header_idx + 1
    if nxt < len(rows) and _is_subheader(rows[nxt]):
        sub = rows[nxt] + [""] * max(0, len(main) - len(rows[nxt]))
        main = [f"{m} > {s}".strip(" >") if s else m for m, s in zip(main, sub)]

    seen: dict[str, int] = {}
    out: list[str] = []
    for h in main:
        if not h:
            out.append("")
            continue
        if h in seen:
            seen[h] += 1
            out.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 1
            out.append(h)
    return out


def _used_columns(
    headers: list[str], rows: list[list[str]], data_start: int
) -> list[int]:
    used = [False] * len(headers)
    for r_idx in range(data_start, len(rows)):
        for c, v in enumerate(rows[r_idx]):
            if c < len(headers) and v:
                used[c] = True
    keep = [c for c in range(len(headers)) if used[c] or headers[c]]
    return keep if keep else list(range(len(headers)))
